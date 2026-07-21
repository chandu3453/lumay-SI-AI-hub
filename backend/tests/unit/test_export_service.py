"""Export service unit tests — CSV/XLSX formula-injection guard (Sprint 30)."""

import csv
import io

from openpyxl import load_workbook

from domains.reporting.services.export_service import rows_to_csv, rows_to_xlsx


class TestFormulaInjectionGuard:
    def test_csv_prefixes_leading_equals_sign(self) -> None:
        csv_bytes = rows_to_csv([{"name": "=cmd|'/c calc'!A1", "count": 1}])
        reader = csv.DictReader(io.StringIO(csv_bytes.decode("utf-8")))
        row = next(reader)
        assert row["name"] == "'=cmd|'/c calc'!A1"

    def test_csv_prefixes_plus_minus_at_triggers(self) -> None:
        csv_bytes = rows_to_csv(
            [
                {"name": "+1+1"},
                {"name": "-1+1"},
                {"name": "@SUM(A1)"},
            ]
        )
        reader = csv.DictReader(io.StringIO(csv_bytes.decode("utf-8")))
        rows = list(reader)
        assert rows[0]["name"] == "'+1+1"
        assert rows[1]["name"] == "'-1+1"
        assert rows[2]["name"] == "'@SUM(A1)"

    def test_csv_leaves_ordinary_values_untouched(self) -> None:
        csv_bytes = rows_to_csv([{"name": "Ahmed Al Badi", "count": 5}])
        reader = csv.DictReader(io.StringIO(csv_bytes.decode("utf-8")))
        row = next(reader)
        assert row["name"] == "Ahmed Al Badi"
        assert row["count"] == "5"

    def test_xlsx_prefixes_leading_equals_sign(self) -> None:
        xlsx_bytes = rows_to_xlsx([{"name": "=HYPERLINK(\"http://evil\")"}])
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet = workbook.active
        assert sheet.cell(row=2, column=1).value == "'=HYPERLINK(\"http://evil\")"

    def test_xlsx_leaves_ordinary_values_untouched(self) -> None:
        xlsx_bytes = rows_to_xlsx([{"name": "Ahmed Al Badi", "count": 5}])
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet = workbook.active
        assert sheet.cell(row=2, column=1).value == "Ahmed Al Badi"
        assert sheet.cell(row=2, column=2).value == 5
